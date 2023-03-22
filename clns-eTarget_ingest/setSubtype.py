 # 
 # This file is part of the eTarget ingest distribution (https://github.com/digital-ECMT/eTarget_ingest).
 # Copyright (C) 2017 - 2021 digital ECMT
 # 
 # This program is free software: you can redistribute it and/or modify  
 # it under the terms of the GNU General Public License as published by  
 # the Free Software Foundation, version 3.
 #
 # This program is distributed in the hope that it will be useful, but 
 # WITHOUT ANY WARRANTY; without even the implied warranty of 
 # MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU 
 # General Public License for more details.
 #
 # You should have received a copy of the GNU General Public License 
 # along with this program. If not, see <http://www.gnu.org/licenses/>.
 #

from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import pymssql

class SetSubtype:
    def __init__(self, remotepassword, 
                 filename="SubtypesForPatients.xlsx", 
                 remotehostname="uomdevsqlserveret.database.windows.net", 
                 remoteusername="uomadmin@uomdevsqlserveret.database.windows.net", 
                 remotedbname="uom_dev_etarget_db"):
        self.excelFile=filename
        self.host=remotehostname
        self.user=remoteusername
        self.pw=remotepassword
        self.db=remotedbname
         
        self.conn = pymssql.connect(self.host,self.user, self.pw, self.db, autocommit=False)
        self.wb = load_workbook(self.excelFile)
        print(self.wb.sheetnames)
        self.ws=self.wb["Sheet1"]

    def process(self):
        for row in self.ws.iter_rows(min_row=2):
            tarid=row[0].value
            category=row[1].value
            subtype=row[2].value
            addInfo=row[3].value
            
            print(str(tarid)+ "; " + str(category)+"; "+ str(subtype)+"; "+str(addInfo))
            self.checkAddSubtype(str(subtype))
            self.updateDB(str(tarid), str(category), str(subtype), str(addInfo))
    
        self.conn.commit()
    
    
    def updateDB(self, tarid,category,subtype,addInfo):
        update="UPDATE CONDITION_OCCURRENCE SET condition_concept_id=(SELECT TOP 1 condition_concept_id FROM CONCEPT_CONDITION_SUBTYPE WHERE subtype_name = '"+subtype+"'), condition_subtype_concept_id=(SELECT TOP 1 subtype_concept_id FROM CONCEPT_CONDITION_SUBTYPE WHERE subtype_name = '"+subtype+"'), additional_details='"+addInfo+"'WHERE person_id=(SELECT person_id from PERSON where target_id='"+tarid+"')"
        cursor=self.conn.cursor()
        cursor.execute(update)

        
        
    def checkAddSubtype(self, subtype):
        select="IF NOT EXISTS (SELECT * FROM CONCEPT_CONDITION_SUBTYPE WHERE subtype_name = '"+subtype+"') INSERT INTO CONCEPT_CONDITION_SUBTYPE (subtype_name, condition_concept_id) values ('"+subtype+"',(select top 1 condition_concept_id from CONCEPT_CONDITION where condition_name='Other'))"
        cursor=self.conn.cursor()
        cursor.execute(select)
        
      
if __name__ == '__main__':
    print(sys.argv[1:])
    print(*sys.argv[1:])
    classinst = SetSubtype(*sys.argv[1:])
    classinst.process()