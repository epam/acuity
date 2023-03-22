 # 
 # This file is part of the eTarget ingest distribution (https://github.com/digital-ECMT/eTarget_ingest).
 # Copyright (C) 2017 - 2021 digital ECMT
 # 
 # This program is free software: you can redistribute it and/or modify  
 # it under the terms of the GNU General Public License as published by  
 # the Free Software Foundation, version 3.
 #
 # This program is distributed in the hope that it will be useful, but 
 # WITHOUT ANY WARRANTY; without even the implied warranty of 
 # MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU 
 # General Public License for more details.
 #
 # You should have received a copy of the GNU General Public License 
 # along with this program. If not, see <http://www.gnu.org/licenses/>.
 #

from openpyxl import Workbook
from openpyxl import load_workbook
from azure.storage.file import FileService
from pycel import ExcelCompiler
from datetime import datetime
import sys
import pymssql
import utilities
import io
import re

class IHC_Report:
    def __init__(self, 
                 filename,
                 remotehostname, 
                 remoteusername, 
                 remotepassword,
                 remotedbname,
                 fileuser, 
                 filekey,
                 datadir='data',
                 logblob='log'):
        self.excelFile=filename
        self.host=remotehostname
        self.user=remoteusername
        self.pw=remotepassword
        self.db=remotedbname
        self.datadir=datadir
        self.file_service = FileService(account_name=fileuser, account_key=filekey)
        output_stream = io.BytesIO()
        self.file_service.get_file_to_stream(self.datadir, None, self.excelFile, output_stream)
        self.wb = load_workbook(output_stream)
        self.ws=self.wb["Report"]
        self.excel = ExcelCompiler(excel=self.wb)
        self.log = utilities.Util(remotehostname, remoteusername, remotepassword, remotedbname, fileuser, filekey, logblob)
        self.conn = pymssql.connect(self.host,self.user, self.pw, self.db, autocommit=False)
    
    def __del__(self):
        self.conn.close()
        
    def ingest(self):
        data={}
        target_id = self.ws['B1'].value
        data["person_id"] = self.queryPerson(target_id)
        data["specimen_id"] = self.querySpecimen(self.ws['B2'].value, data["person_id"])
#         if self.checkRecordExists(data["specimen_id"])== True :
#             self.log.logMessage("Record exits for " + str(target_id))
#             self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(),'Warning: report exists')
#             return(None)
        resubmission=self.checkResubmission(data["specimen_id"])
#         if resubmission[0]==2:
#             self.log.logMessage("Resubmission after patient being discussed  " + str(target_id))
#             self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(),'Error: report cannot be overwritten')
#             return(None)
        if resubmission[0]==1:
            self.log.logMessage("Resubmission - overwrite data  " + str(target_id))
            self.deleteRecord(resubmission[1])
        data["date_received"] = self.parseDate(self.ws['B3'].value)
        data["date_report"] = self.parseDate(self.ws['B4'].value)
        data["cd3_area"] = self.parseFloat(self.excel.evaluate('Report!B5'))
        data["cd3_tumoural"] = self.parseFloat(self.excel.evaluate('Report!B6'))
        data["cd3_stromal"] = self.parseFloat(self.excel.evaluate('Report!B7'))
        data["cd8_area"] = self.parseFloat(self.excel.evaluate('Report!B8'))
        data["cd8_tumoural"] = self.parseFloat(self.excel.evaluate('Report!B9'))
        data["cd8_stromal"] = self.parseFloat(self.excel.evaluate('Report!B10'))
        data["pdl1_tps"] = self.ws['B11'].value
        data["estimated_result"] = self.ws['B12'].value
        data["comments"] = self.ws['B13'].value
        print(str(data))
        id=self.insertIHCRecord(data);
        if id>0:
            self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(),'Success')
            self.commit()
            return(data)
        self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(),'Error: Database commit not successful')
        return(None)
        
        
    def queryPerson(self, target_id):
        if len(target_id)==0 or len(target_id)>10:
            self.log.logMessage("Patient id of wrong size " + str(target_id))
            self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(),'Error: TargetID has wrong size.')
            raise Exception('Patient ID is empty or too long')
        query="SELECT person_id FROM PERSON where target_id='"+str(target_id)+"'"
        cursor=self.conn.cursor()
        cursor.execute(query)
        row=cursor.fetchone()
        if row is None:
            self.log.logMessage("Can't find Patient id  " + str(target_id))
            self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(),'Error: Patient ID not found in database.')
            raise Exception('Cannot find Target ID')
        return(row[0])
       
    def querySpecimen(self, specimentext, person_id):
        if len(specimentext)== 0:
            self.log.logMessage("SampleID is empty ")
            self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(),'Error: SampleID empty')
            raise Exception('Sample ID is empty')
        
        pattern = re.compile("^[A-Z]{3}\d{7}Bx\d{1,2}IC")
        if not re.match(pattern, specimentext):
            self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(), 'Error: specimen ID is incompatible with the naming convention')
            self.log.logMessage('Error ingesting IHC data - sample ID has the wrong structure')
            raise Exception('Error sample id does not match pattern')
        timepoint=specimentext[10:]
        
#         pattern = re.compile("^T\d{1,2}(\D|$)")
#         if not re.match(pattern, timepoint): 
#             self.log.systemStatusUpdate(self.filename, 'IHC', self.log.timestamp(), 'Error: specimen ID is incompatible with the naming convention - timepoint')
#             self.log.logMessage('Error ingesting IHC data - sample ID has the wrong structure - timepoint')
#             raise Exception('Error sample id does not match pattern')
        p2 = re.compile('\d{1,2}')
        t = p2.findall(timepoint)[0]
        query="select specimen_id, SPECIMEN.person_id FROM SPECIMEN left join PERSON on PERSON.person_id=SPECIMEN.person_id where specimen_concept_id!=1 and baseline_number="+t+" and target_id='"+specimentext[:10]+"'"
        print(query)
        cursor= self.conn.cursor()
        cursor.execute(query)
        row=cursor.fetchone()
        if row is None:
            self.log.logMessage("Sample not found " + specimentext)
            self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(),'Error: Sample ID not found')
            raise Exception('Cannot find Specimen ID')
        if row[1]!=person_id:
            self.log.logMessage("Sample ID does not fit patient ")
            self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(),'Error: Sample ID does not fit patient')
            raise Exception('Specimen ID does not fit patient')
        return(row[0])
    
    def parseDate(self, date):
        if date is None or len(date)==0:
            return 'NULL'
        return datetime.strptime(date,'%d%b%Y').strftime('%Y-%m-%d')
    
    def parseFloat(self, float):
        if isinstance(float,str):
            return('NULL')
        if float is None:
            return('NULL')
        return(float)
    
    def insertIHCRecord(self, data):
        insert="INSERT INTO IHC_REPORT (person_id, specimen_id, sample_received_date, report_date, cd3_total_tissue_area, cd3_intratumoural, cd3_intrastromal, cd8_total_tissue_area, cd8_intratumoural, cd8_intrastromal, pdl1_tps, estimated_results,  comments) values("+str(data['person_id'])+", "+str(data['specimen_id'])+", '"+str(data['date_received'])+"', '"+str(data['date_report'])+"', "+str(data['cd3_area'])+"," + str(data['cd3_tumoural'])+", "+str(data['cd3_stromal'])+", "+str(data['cd8_area'])+", "+str(data['cd8_tumoural'])+", "+str(data['cd8_stromal'])+", '"+str(data['pdl1_tps'])+"', '" + str(data['estimated_result'])+"', '"+str(data['comments'])+"');"
        print(insert)
        cursor= self.conn.cursor()
        cursor.execute(insert)
        return cursor.lastrowid
        
    def checkRecordExists(self, specimen_id):
        select="select * from IHC_REPORT where specimen_id="+str(specimen_id)
        cursor=self.conn.cursor()
        cursor.execute(select)
        row=cursor.fetchone()
        if row is None:
            return(False)
        return(True)
    
        # checkResubmission IHC report
    def checkResubmission(self, specimen_id):
        checkSubmission = "select person_id, ingestion_date, ihc_report_id from IHC_REPORT where specimen_id="+str(specimen_id)
        cursorRun = self.conn.cursor()
        cursorRun.execute(checkSubmission)
        row = cursorRun.fetchone()
        cursorRun.close()
        if row is None:
            return 0,0
         
        person_id=None
        if row[0] is not None:
            person_id=row[0]
        ingestion_date = row[1]
        ihc_report_id = row[2]
         
        lastReportSQL = "SELECT created_on FROM MEETING_OUTCOME as mo where mo.person_id="+str(person_id)+" ORDER BY created_on DESC;"
        cursor = self.conn.cursor()
        cursor.execute(lastReportSQL)
        row = cursor.fetchone();
        cursor.close()
        if row is not None:
            lastDiscussedDate = row[0]
            if ingestion_date is not None:
                if lastDiscussedDate>=ingestion_date:
                    return 2,ihc_report_id
                else: 
                    return 1,ihc_report_id
             
        return 1,ihc_report_id

    def deleteRecord(self, ihc_report_id):
        if ihc_report_id is None or type(ihc_report_id) is not int:
            return
        deleteIHCRecord = "delete from IHC_REPORT where ihc_report_id =" + str(ihc_report_id)
        cursor = self.conn.cursor()
        cursor.execute(deleteIHCRecord)
        
        
    def commit(self):
        self.conn.commit();
        
    def deleteFile(self):
        try:
            if self.file_service.exists(self.datadir, None, self.excelFile):
                self.file_service.delete_file(self.datadir, None, self.excelFile)
            self.log.systemStatusUpdate(self.excelFile, 'IHC', self.log.timestamp(), 'Success')
        except:
            self.log.logMessage('There was a problem deleting '+self.excelFile)
